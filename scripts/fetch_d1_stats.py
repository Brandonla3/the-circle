#!/usr/bin/env python3
"""
fetch_d1_stats.py
-----------------
Pulls player stats (batters, pitchers, fielding) for every D1 softball
conference from the Boost Sport API and writes:

  1. data/d1_softball_stats_<YEAR>.xlsx  -- organized Excel workbook
  2. data/json/<conf_slug>/<section>.json -- one JSON file per conference+section
     (these feed the Next.js /api/d1-stats route on Vercel)

JSON shape per file:
    {
        "conference": "Big Ten",
        "section": "batters",
        "season": 2026,
        "fetchedAt": "2026-04-18T05:00:00Z",
        "columns": ["name", "team", "gp", ...],
        "rows": [ [...], [...], ... ]
    }

Run manually:
    pip install requests openpyxl
    python scripts/fetch_d1_stats.py

Called automatically every Friday night by GitHub Actions.
"""

import json
import os
import re
import time

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone

# ---------------------------------------------------------------------------
# Conference configuration
# ---------------------------------------------------------------------------
BOOST_CONFERENCES = [
    {"name": "Big Ten",       "api_name": "Big Ten",              "slug": "big-ten"},
    {"name": "ACC",           "api_name": "ACC",                  "slug": "acc"},
    {"name": "Big 12",        "api_name": "Big 12",               "slug": "big-12"},
    {"name": "SEC",           "api_name": "SEC",                  "slug": "sec"},
    {"name": "Pac-12",        "api_name": "Pac-12",               "slug": "pac-12"},
    {"name": "American",      "api_name": "American Athletic",    "slug": "american"},
    {"name": "Mountain West", "api_name": "Mountain West",        "slug": "mountain-west"},
    {"name": "Sun Belt",      "api_name": "Sun Belt",             "slug": "sun-belt"},
    {"name": "C-USA",         "api_name": "Conference USA",       "slug": "c-usa"},
    {"name": "MAC",           "api_name": "Mid-American",         "slug": "mac"},
    {"name": "MVC",           "api_name": "Missouri Valley",      "slug": "mvc"},
    {"name": "WAC",           "api_name": "Western Athletic",     "slug": "wac"},
    {"name": "Big West",      "api_name": "Big West",             "slug": "big-west"},
    {"name": "Southern",      "api_name": "Southern",             "slug": "southern"},
    {"name": "Southland",     "api_name": "Southland",            "slug": "southland"},
    {"name": "SWAC",          "api_name": "Southwestern Athletic","slug": "swac"},
    {"name": "MEAC",          "api_name": "Mid-Eastern Athletic", "slug": "meac"},
    {"name": "NEC",           "api_name": "Northeast",            "slug": "nec"},
    {"name": "OVC",           "api_name": "Ohio Valley",          "slug": "ovc"},
    {"name": "Patriot",       "api_name": "Patriot",              "slug": "patriot"},
    {"name": "MAAC",          "api_name": "Metro Atlantic Athletic","slug": "maac"},
    {"name": "A-10",          "api_name": "Atlantic 10",          "slug": "a-10"},
    {"name": "Big South",     "api_name": "Big South",            "slug": "big-south"},
    {"name": "CAA",           "api_name": "Coastal Athletic",     "slug": "caa"},
    {"name": "Horizon",       "api_name": "Horizon",              "slug": "horizon"},
    {"name": "Ivy League",    "api_name": "Ivy",                  "slug": "ivy"},
    {"name": "Summit",        "api_name": "Summit",               "slug": "summit"},
    {"name": "WCC",           "api_name": "West Coast",           "slug": "wcc"},
    {"name": "ASUN",          "api_name": "ASUN",                 "slug": "asun"},
    {"name": "America East",  "api_name": "America East",         "slug": "america-east"},
]

STAT_SECTIONS = ["batters", "pitchers", "fielding"]

BOOST_BASE = (
    "https://engage-api.boostsport.ai/api/sport/sb/stats/table"
    "?conference={conf}"
    "&seasons={season}"
    "&view=table"
    "&type=player"
    "&split=all"
    "&teams=all"
    "&section={section}"
    "&level=season"
    "&limit=2000"
    "&orderBy=default_rank"
    "&order=asc"
)

REQ_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; the-circle-stats-bot/1.0)",
    "Accept": "application/json",
}

SEASON = datetime.now(timezone.utc).year

# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
NORM_FILL   = PatternFill("solid", fgColor="FFFFFF")


def style_header_row(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_size_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)


def add_zebra_rows(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if r % 2 == 0 else NORM_FILL
        for c in range(1, col_count + 1):
            ws.cell(row=r, column=c).fill = fill


# ---------------------------------------------------------------------------
# Data fetching & parsing
# ---------------------------------------------------------------------------
def fetch_boost(conf_api_name, section, season=SEASON, retries=3):
    url = BOOST_BASE.format(
        conf=requests.utils.quote(conf_api_name),
        season=season,
        section=section,
    )
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=REQ_HEADERS, timeout=20)
            if resp.status_code == 200:
                return resp.json()
            if resp.status_code == 404:
                return None
            print(f"  [{section}] HTTP {resp.status_code} for {conf_api_name}, attempt {attempt+1}")
        except requests.RequestException as exc:
            print(f"  [{section}] Error for {conf_api_name}: {exc}, attempt {attempt+1}")
        time.sleep(2 ** attempt)
    return None


def flatten_record(rec):
    """
    Flatten one API record dict.  Nested sub-dicts (e.g. player: {name, team})
    get merged into the top level; sub-keys that collide keep the parent key name.
    """
    flat = {}
    for k, v in rec.items():
        if isinstance(v, dict):
            for sub_k, sub_v in v.items():
                flat_key = sub_k if sub_k not in rec else f"{k}_{sub_k}"
                flat[flat_key] = sub_v
        else:
            flat[k] = v
    return flat


def extract_rows(payload):
    """
    Parse the Boost API response into (columns_list, rows_list).
    Returns ([], []) when there's no usable data.
    """
    if not payload:
        return [], []

    data = payload
    if isinstance(data, dict):
        for key in ("data", "players", "stats", "results"):
            if key in data and data[key]:
                data = data[key]
                break

    if not isinstance(data, list) or len(data) == 0:
        return [], []

    flat_records = [flatten_record(r) for r in data]
    columns = list(flat_records[0].keys())
    rows = [[r.get(c, "") for c in columns] for r in flat_records]
    return columns, rows


# ---------------------------------------------------------------------------
# JSON output helpers
# ---------------------------------------------------------------------------
def write_json(conf_slug, section, columns, rows, fetched_at):
    out_dir = os.path.join("data", "json", conf_slug)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{section}.json")

    # Build list-of-objects for easy consumption in JS
    records = [dict(zip(columns, row)) for row in rows]
    payload = {
        "conference": conf_slug,
        "section": section,
        "season": SEASON,
        "fetchedAt": fetched_at,
        "columns": columns,
        "count": len(rows),
        "players": records,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    return out_path


def write_index(all_meta):
    """Write data/json/index.json -- a lightweight manifest of what's available."""
    out_path = os.path.join("data", "json", "index.json")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({
            "season": SEASON,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "conferences": all_meta,
        }, f, ensure_ascii=False, indent=2)
    print(f"  -> index written: {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    os.makedirs("data", exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws.append(["Conference", "Category", "Players", "Fetched At"])
    style_header_row(summary_ws, 4)
    summary_ws.row_dimensions[1].height = 20
    summary_row = 2

    all_meta = []
    total_players = 0
    fetched_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    for conf in BOOST_CONFERENCES:
        conf_display = conf["name"]
        conf_api = conf["api_name"]
        conf_slug = conf["slug"]
        conf_meta = {"name": conf_display, "slug": conf_slug, "sections": {}}

        print(f"\n{'='*55}")
        print(f"Conference: {conf_display}")

        for section in STAT_SECTIONS:
            category_display = section.capitalize()
            print(f"  {category_display}...", end=" ", flush=True)

            payload = fetch_boost(conf_api, section)
            columns, rows = extract_rows(payload)

            # -- Excel sheet --
            sheet_name = f"{conf_display} - {category_display}"[:31]
            ws = wb.create_sheet(sheet_name)

            if not columns:
                ws.append(["No data returned."])
                print("no data")
                conf_meta["sections"][section] = 0
            else:
                ws.append(columns)
                style_header_row(ws, len(columns))
                ws.row_dimensions[1].height = 20

                for row in rows:
                    ws.append(row)

                add_zebra_rows(ws, 2, len(rows) + 1, len(columns))
                auto_size_columns(ws)
                ws.freeze_panes = "A2"

                print(f"{len(rows)} players", end="")

                # -- JSON file --
                json_path = write_json(conf_slug, section, columns, rows, fetched_at)
                print(f" -> {json_path}")

                total_players += len(rows)
                conf_meta["sections"][section] = len(rows)

            summary_ws.cell(summary_row, 1, conf_display)
            summary_ws.cell(summary_row, 2, category_display)
            summary_ws.cell(summary_row, 3, len(rows))
            summary_ws.cell(summary_row, 4, fetched_at)
            summary_row += 1

            time.sleep(0.4)

        all_meta.append(conf_meta)

    # Finalise Excel summary sheet
    auto_size_columns(summary_ws)
    add_zebra_rows(summary_ws, 2, summary_row - 1, 4)
    summary_ws.freeze_panes = "A2"

    xlsx_path = os.path.join("data", f"d1_softball_stats_{SEASON}.xlsx")
    wb.save(xlsx_path)

    # Write JSON index manifest
    write_index(all_meta)

    print(f"\n{'='*55}")
    print(f"Excel saved : {xlsx_path}")
    print(f"JSON dir    : data/json/")
    print(f"Total rows  : {total_players}")
    print(f"Sheets      : {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
